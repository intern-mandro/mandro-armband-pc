"""
analyze_accuracy_by_arm.py
==========================
Plot model accuracy vs subject arm / armband size, controlling for the
gesture set and excluding failed trainings.

Reads accuracies from sessions.xlsx and joins them by subject_id with the
anthropometric table below (edit SUBJECTS to your recruitment sheet).

USAGE (project root, app closed):
    python analyze_accuracy_by_arm.py                 # 6cl, offline
    python analyze_accuracy_by_arm.py --set rps
    python analyze_accuracy_by_arm.py --metric both   # offline + online
    python analyze_accuracy_by_arm.py --xlsx client_app/sessions.xlsx

Output: accuracy_by_arm.png + a per-subject summary printed to the console.

HONEST CAVEAT (keep in your report): armband is chosen from arm size, so the
two are confounded; N is small; subjects were recorded at different periods.
Exploratory only - not a causal result.
"""
import os
import sys
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

# ─── EDIT ME: anthropometrics per subject (cm). arm=None -> not plotted. ──
SUBJECTS = {
    "S001": {"arm": 27, "armband": 21, "sex": "M"},
    "S002": {"arm": 21, "armband": 21, "sex": "F"},
    "S003": {"arm": 30, "armband": 21, "sex": "F"},
    "S004": {"arm": 29, "armband": 21, "sex": "M"},
    "S005": {"arm": 21, "armband": 21, "sex": "F"},
    "S006": {"arm": 21, "armband": 21, "sex": "M"},   # verify arm
    "S008": {"arm": 22, "armband": 19, "sex": "M"},
    "S011": {"arm": 20, "armband": 19, "sex": "F"},
    "S012": {"arm": 19, "armband": 19, "sex": "F"},
    "S013": {"arm": None, "armband": None, "sex": "?"},  # <- fill from sheet
}

MIN_VALID_ACC = 0.5          # below this = failed/aborted training, excluded
ARMBAND_COLORS = {19: "#3aa0ff", 21: "#a06bd6"}
SEX_MARKERS = {"M": "s", "F": "o"}


def find_xlsx(explicit):
    if explicit:
        return explicit if os.path.isfile(explicit) else None
    for c in ("client_app/sessions.xlsx", "sessions.xlsx",
              os.path.join("..", "client_app", "sessions.xlsx")):
        if os.path.isfile(c):
            return c
    return None


def resolve_set(row, idx):
    gs = row[idx["gesture_set"]] if "gesture_set" in idx else None
    if gs:
        return str(gs).strip().lower()
    mp = str(row[idx["model_path"]] or "").lower()
    for t in ("rps", "6cl", "4cl"):
        if f"_{t}_" in mp or f"_{t}." in mp:
            return t
    return {6: "6cl", 4: "4cl"}.get(row[idx["n_classes"]])


def load(xlsx, want_set):
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    agg = {}          # subject -> {"off":[...], "on":[...]}
    excluded = []     # (subject, set, acc) failed runs
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = row[idx["subject_id"]]
        if not sid:
            continue
        sid = str(sid).strip()
        if resolve_set(row, idx) != want_set:
            continue
        d = agg.setdefault(sid, {"off": [], "on": []})
        off, on = row[idx["acc_offline"]], row[idx["acc_online"]]
        if isinstance(off, (int, float)):
            if off < MIN_VALID_ACC:
                excluded.append((sid, want_set, off))
            else:
                d["off"].append(off * 100.0)
        if isinstance(on, (int, float)) and on >= MIN_VALID_ACC:
            d["on"].append(on * 100.0)
    return agg, excluded


def stats(vals):
    return (sum(vals) / len(vals), min(vals), max(vals), len(vals)) if vals else None


def points(agg, metric):
    key = "off" if metric == "offline" else "on"
    out = []
    for sid, info in SUBJECTS.items():
        if info.get("arm") is None:
            continue
        st = stats(agg.get(sid, {}).get(key, []))
        if not st:
            continue
        mean, lo, hi, n = st
        out.append({"sid": sid, "arm": info["arm"], "armband": info["armband"],
                    "sex": info.get("sex", "?"), "fit": info["arm"] - info["armband"],
                    "acc": mean, "lo": lo, "hi": hi, "n": n})
    return out


def scatter(ax, pts, xkey, xlabel, title):
    for p in pts:
        ax.errorbar(p[xkey], p["acc"], yerr=[[p["acc"] - p["lo"]], [p["hi"] - p["acc"]]],
                    fmt="none", ecolor="#aaa", elinewidth=1, capsize=3, zorder=2)
        ax.scatter(p[xkey], p["acc"], c=ARMBAND_COLORS.get(p["armband"], "#888"),
                   marker=SEX_MARKERS.get(p["sex"], "o"), s=130,
                   edgecolors="#222", linewidths=0.6, zorder=3)
        ax.annotate(f"{p['sid']} (n={p['n']})", (p[xkey], p["acc"]),
                    textcoords="offset points", xytext=(7, 4), fontsize=8)
    ax.set_xlabel(xlabel); ax.set_ylabel("Accuracy (%)")
    ax.set_title(title, fontsize=11, fontweight="bold"); ax.grid(alpha=0.25)


def strip(ax, pts):
    groups = {}
    for p in pts:
        groups.setdefault(p["armband"], []).append(p)
    xs = sorted(groups)
    for i, band in enumerate(xs):
        vals = groups[band]
        for j, p in enumerate(vals):
            ax.scatter(i + (j - (len(vals) - 1) / 2) * 0.08, p["acc"],
                       c=ARMBAND_COLORS.get(band, "#888"),
                       marker=SEX_MARKERS.get(p["sex"], "o"), s=130,
                       edgecolors="#222", linewidths=0.6, zorder=3)
        m = sum(p["acc"] for p in vals) / len(vals)
        ax.plot([i - 0.2, i + 0.2], [m, m], color="#222", lw=2)
        ax.annotate(f"mean {m:.1f}%", (i, m), textcoords="offset points",
                    xytext=(14, -2), fontsize=8)
    ax.set_xticks(range(len(xs))); ax.set_xticklabels([f"{b} cm" for b in xs])
    ax.set_xlabel("Armband size"); ax.set_ylabel("Accuracy (%)")
    ax.set_title("Accuracy by armband size", fontsize=11, fontweight="bold")
    ax.grid(alpha=0.25, axis="y")


def barplot(ax, agg, metrics):
    rows = []
    for sid, info in SUBJECTS.items():
        if info.get("arm") is None:
            continue
        so = stats(agg.get(sid, {}).get("off", []))
        sn = stats(agg.get(sid, {}).get("on", []))
        if not so and not sn:
            continue
        rows.append({"sid": sid, "armband": info["armband"], "arm": info["arm"],
                     "off": so, "on": sn})
    rows.sort(key=lambda r: (r["armband"], r["arm"]))

    want_on = "online" in metrics
    width = 0.38 if want_on else 0.6
    for i, r in enumerate(rows):
        col = ARMBAND_COLORS.get(r["armband"], "#888")
        if r["off"]:
            m, lo, hi, n = r["off"]
            xo = i - (width / 2 if want_on else 0)
            ax.bar(xo, m, width, color=col, edgecolor="#222", linewidth=0.6, zorder=2)
            ax.errorbar(xo, m, yerr=[[m - lo], [hi - m]], fmt="none",
                        ecolor="#333", elinewidth=1, capsize=3, zorder=3)
            ax.annotate(f"{m:.0f}\n(n={n})", (xo, m), ha="center", va="bottom",
                        textcoords="offset points", xytext=(0, 2), fontsize=7)
        if want_on and r["on"]:
            m, lo, hi, n = r["on"]
            xn = i + width / 2
            ax.bar(xn, m, width, color=col, edgecolor="#222", linewidth=0.6,
                   hatch="//", alpha=0.85, zorder=2)
            ax.errorbar(xn, m, yerr=[[m - lo], [hi - m]], fmt="none",
                        ecolor="#333", elinewidth=1, capsize=3, zorder=3)
            ax.annotate(f"{m:.0f}\n(n={n})", (xn, m), ha="center", va="bottom",
                        textcoords="offset points", xytext=(0, 2), fontsize=7)

    # separator + group mean (offline) between armband groups
    prev = None
    for i, r in enumerate(rows):
        if prev is not None and r["armband"] != prev:
            ax.axvline(i - 0.5, color="#ccc", ls="--", lw=1)
        prev = r["armband"]
    for band in sorted({r["armband"] for r in rows}):
        offs = [r["off"][0] for r in rows if r["armband"] == band and r["off"]]
        if offs:
            xs = [i for i, r in enumerate(rows) if r["armband"] == band]
            ax.plot([min(xs) - 0.4, max(xs) + 0.4], [sum(offs) / len(offs)] * 2,
                    color=ARMBAND_COLORS.get(band, "#888"), lw=2, ls=":")

    ax.set_xticks(range(len(rows)))
    ax.set_xticklabels([f"{r['sid']}\n{r['arm']}cm" for r in rows], fontsize=8)
    ax.set_ylabel("Accuracy (%)")
    ax.set_ylim(0, 105)
    ax.grid(alpha=0.25, axis="y")
    handles = [plt.Rectangle((0, 0), 1, 1, facecolor=c, edgecolor="#222",
                             label=f"armband {b} cm")
               for b, c in sorted(ARMBAND_COLORS.items())]
    handles.append(plt.Rectangle((0, 0), 1, 1, facecolor="#bbb", edgecolor="#222",
                                 label="offline (solid)"))
    if want_on:
        handles.append(plt.Rectangle((0, 0), 1, 1, facecolor="#bbb", edgecolor="#222",
                                      hatch="//", label="online (hatched)"))
    ax.legend(handles=handles, fontsize=8, ncol=2, loc="lower right")


def main():
    explicit = sys.argv[sys.argv.index("--xlsx") + 1] if "--xlsx" in sys.argv else None
    want_set = sys.argv[sys.argv.index("--set") + 1] if "--set" in sys.argv else "6cl"
    metric = sys.argv[sys.argv.index("--metric") + 1] if "--metric" in sys.argv else "offline"

    xlsx = find_xlsx(explicit)
    if not xlsx:
        raise SystemExit("sessions.xlsx not found. Pass --xlsx <path>.")
    print(f"Reading: {xlsx}   (set={want_set})")
    agg, excluded = load(xlsx, want_set)
    if excluded:
        print(f"Excluded {len(excluded)} failed run(s) (acc < {MIN_VALID_ACC}):")
        for sid, s, a in excluded:
            print(f"   {sid} {s}: {a:.3f}")

    print("\nPer-subject summary (set =", want_set, "):")
    print(f"  {'subj':5} {'arm':>5} {'band':>4} {'off_mean':>8} {'n_off':>5} {'on_mean':>8} {'n_on':>4}")
    for sid, info in SUBJECTS.items():
        a = agg.get(sid)
        if not a:
            continue
        so, sn = stats(a["off"]), stats(a["on"])
        print(f"  {sid:5} {str(info['arm']):>5} {str(info['armband']):>4} "
              f"{(f'{so[0]:.1f}' if so else '-'):>8} {(so[3] if so else 0):>5} "
              f"{(f'{sn[0]:.1f}' if sn else '-'):>8} {(sn[3] if sn else 0):>4}")

    plot = sys.argv[sys.argv.index("--plot") + 1] if "--plot" in sys.argv else "scatter"
    metrics = ["offline", "online"] if metric == "both" else [metric]
    metrics = [m for m in metrics if points(agg, m)]
    if not metrics:
        raise SystemExit(f"\nNo '{metric}' data for set={want_set}.")

    if plot == "bar":
        fig, ax = plt.subplots(figsize=(max(9, 1.3 * len(SUBJECTS)), 6))
        barplot(ax, agg, metrics)
        ax.set_title(f"Model accuracy per subject  (set={want_set}, failed runs excluded)",
                     fontsize=13, fontweight="bold")
        fig.text(0.5, 0.005, "Exploratory: armband chosen from arm (confounded); small N; "
                 "different recording periods. Error bars = min-max. Not causal.",
                 ha="center", fontsize=8, color="#555")
        fig.tight_layout(rect=[0, 0.03, 1, 1])
        out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "accuracy_bar.png")
        fig.savefig(out, dpi=150)
        print(f"\nSaved: {out}")
        return

    fig, axes = plt.subplots(len(metrics), 3, figsize=(15, 4.4 * len(metrics)), squeeze=False)
    for r, m in enumerate(metrics):
        pts = points(agg, m)
        scatter(axes[r][0], pts, "arm", "arm size (cm)", f"{m.capitalize()} vs arm  (set={want_set})")
        scatter(axes[r][1], pts, "fit", "arm - armband (cm)  [+ = tighter]", f"{m.capitalize()} vs fit")
        strip(axes[r][2], pts)
        if r == 0:
            h = [plt.Line2D([0], [0], marker="o", color="w", markerfacecolor=c,
                            markeredgecolor="#222", markersize=10, label=f"armband {b} cm")
                 for b, c in sorted(ARMBAND_COLORS.items())]
            h += [plt.Line2D([0], [0], marker=mk, color="w", markerfacecolor="#bbb",
                             markeredgecolor="#222", markersize=10, label=f"sex {sx}")
                  for sx, mk in SEX_MARKERS.items()]
            axes[r][0].legend(handles=h, fontsize=8, loc="best")

    fig.suptitle(f"Model accuracy vs arm / armband size  (set={want_set}, failed runs excluded)",
                 fontsize=14, fontweight="bold")
    fig.text(0.5, 0.005, "Exploratory: armband chosen from arm (confounded); small N; "
             "different recording periods. Error bars = min-max across a subject's sessions. Not causal.",
             ha="center", fontsize=8, color="#555")
    fig.tight_layout(rect=[0, 0.03, 1, 0.96])
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "accuracy_by_arm.png")
    fig.savefig(out, dpi=150)
    print(f"\nSaved: {out}")


if __name__ == "__main__":
    main()