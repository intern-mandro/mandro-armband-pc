"""
subjects_registry.py
====================
Small pure-Python module that owns the lifecycle of `subjects.xlsx`.

The Excel file lives next to this module (`client_app/subjects.xlsx`) and
contains three columns: subject_id, name, surname.

This module is intentionally Qt-free so it can be unit-tested from the
command line:

    python -c "from client_app.subjects_registry import next_id; print(next_id())"

ID format: 'S001', 'S002', ... — 3-digit zero-padded, prefixed with 'S'.
Reasoning: zero-padding keeps alphanumeric sort = numeric sort; the 'S'
prefix avoids collisions with other ID kinds we may introduce later.

Names and surnames are stored UPPERCASE for visual consistency with the
data/ folder naming convention.
"""

from __future__ import annotations
import os
import re
from typing import Optional

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
SUBJECTS_XLSX = os.path.join(_HERE, "subjects.xlsx")

ID_RE = re.compile(r"^S(\d{3})$")
EXPECTED_HEADERS = ["subject_id", "name", "surname"]


# ---------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------
def _load_workbook() -> openpyxl.Workbook:
    """Load subjects.xlsx, validating it has the expected headers."""
    if not os.path.exists(SUBJECTS_XLSX):
        # Auto-create an empty registry. subjects.xlsx is git-ignored (it may
        # hold personal data), so a fresh clone will not have it.
        wb = openpyxl.Workbook()
        wb.active.append(EXPECTED_HEADERS)
        wb.save(SUBJECTS_XLSX)
        return wb
    wb = openpyxl.load_workbook(SUBJECTS_XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise ValueError(
            f"Unexpected headers in {SUBJECTS_XLSX}: {headers}. "
            f"Expected first columns to be {EXPECTED_HEADERS}."
        )
    return wb


def _id_to_int(subject_id: str) -> Optional[int]:
    """Parse 'S001' -> 1; return None for anything else."""
    m = ID_RE.match(subject_id or "")
    return int(m.group(1)) if m else None


def _int_to_id(n: int) -> str:
    """1 -> 'S001'."""
    return f"S{n:03d}"


# ---------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------
def next_id() -> str:
    """Compute the next available subject ID.

    Strategy: read all existing IDs, parse the numeric part, return
    max + 1. If the file is empty, return 'S001'.
    """
    wb = _load_workbook()
    ws = wb.active
    max_n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        n = _id_to_int(str(row[0]).strip())
        if n is not None and n > max_n:
            max_n = n
    return _int_to_id(max_n + 1)


def add_subject(name: str = "", surname: str = "") -> str:
    """Append a new (anonymous) subject row, return the freshly generated ID.

    Names are no longer collected for anonymization. The name/surname
    columns are kept (written empty) so subjects.xlsx keeps its schema.
    """
    name = (name or "").strip().upper()
    surname = (surname or "").strip().upper()

    new_id = next_id()

    wb = _load_workbook()
    ws = wb.active
    ws.append([new_id, name, surname])
    wb.save(SUBJECTS_XLSX)
    return new_id


def list_subjects() -> list[dict]:
    """Return all subjects as a list of dicts:
        [{"subject_id": "S001", "name": "KOTA", "surname": "YAMAMOTO"}, ...]
    """
    wb = _load_workbook()
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        out.append({
            "subject_id": str(row[0]).strip(),
            "name":       str(row[1] or "").strip(),
            "surname":    str(row[2] or "").strip(),
        })
    return out


def find_subject(subject_id: str) -> Optional[dict]:
    """Lookup a subject by ID. Returns None if not found."""
    for s in list_subjects():
        if s["subject_id"] == subject_id:
            return s
    return None


def delete_subjects(subject_ids) -> int:
    """Delete every subject whose subject_id is in `subject_ids`.

    Accepts a single id (str) or an iterable of ids. Returns the number
    of rows actually removed.

    Note: this only touches subjects.xlsx. Sessions recorded for these
    subjects in sessions.xlsx are left untouched (they become orphan
    rows referencing a missing subject_id).
    """
    if isinstance(subject_ids, str):
        subject_ids = [subject_ids]
    targets = {str(s).strip() for s in subject_ids if str(s).strip()}
    if not targets:
        return 0

    wb = _load_workbook()
    ws = wb.active

    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        if str(row[0]).strip() in targets:
            rows_to_delete.append(i)

    for i in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(i, 1)

    if rows_to_delete:
        wb.save(SUBJECTS_XLSX)
    return len(rows_to_delete)


def display_label(subject: dict) -> str:
    """Format a subject for the dropdown. Names are not shown (anonymization)."""
    return f"{subject['subject_id']}"


# ---------------------------------------------------------------------
# CLI smoke test
# ---------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    cmd = sys.argv[1] if len(sys.argv) > 1 else "list"

    if cmd == "next":
        print(next_id())

    elif cmd == "add":
        if len(sys.argv) < 4:
            print("usage: subjects_registry.py add <name> <surname>")
            sys.exit(2)
        sid = add_subject(sys.argv[2], sys.argv[3])
        print(f"added: {sid}")

    elif cmd == "list":
        for s in list_subjects():
            print(display_label(s))

    elif cmd == "find":
        if len(sys.argv) < 3:
            print("usage: subjects_registry.py find <subject_id>")
            sys.exit(2)
        s = find_subject(sys.argv[2])
        print(s if s else "not found")

    elif cmd == "delete":
        if len(sys.argv) < 3:
            print("usage: subjects_registry.py delete <subject_id> [<subject_id> ...]")
            sys.exit(2)
        n = delete_subjects(sys.argv[2:])
        print(f"deleted {n} subject(s)")

    else:
        print(f"unknown command: {cmd}")
        print("commands: list | next | add <name> <surname> | find <id> "
              "| delete <id> [<id> ...]")
        sys.exit(2)