"""
sessions_registry.py
====================
Small pure-Python module that owns the lifecycle of `sessions.xlsx`.

The Excel file lives next to this module (`client_app/sessions.xlsx`).
Columns:
    subject_id, date, hour, n_classes,
    takes_batch1, takes_batch2,
    acc_offline, f1_offline,
    acc_online, f1_online,
    model_path, source_log, data_folder

Like subjects_registry, this module is intentionally Qt-free.

Workflow:
- Phase 2 (training) calls add_session() at the end with the offline
  metrics it just measured.
- Phase 5 (online scoring) calls update_online_metrics() to fill the
  acc_online / f1_online columns of an existing row.
"""

from __future__ import annotations
import os
from datetime import datetime
from typing import Optional

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
SESSIONS_XLSX = os.path.join(_HERE, "sessions.xlsx")

EXPECTED_HEADERS = [
    "subject_id", "date", "hour", "n_classes",
    "takes_batch1", "takes_batch2",
    "acc_offline", "f1_offline",
    "acc_online", "f1_online",
    "model_path", "source_log", "data_folder",
]

# Column index (1-based for openpyxl, matches order in EXPECTED_HEADERS)
COL = {h: i + 1 for i, h in enumerate(EXPECTED_HEADERS)}


# ---------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------
def _load_workbook() -> openpyxl.Workbook:
    if not os.path.exists(SESSIONS_XLSX):
        # Auto-create an empty registry. sessions.xlsx is git-ignored (data is
        # local to each machine), so a fresh clone will not have it.
        wb = openpyxl.Workbook()
        wb.active.append(EXPECTED_HEADERS)
        wb.save(SESSIONS_XLSX)
        return wb
    wb = openpyxl.load_workbook(SESSIONS_XLSX)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise ValueError(
            f"Unexpected headers in {SESSIONS_XLSX}: {headers}. "
            f"Expected: {EXPECTED_HEADERS}."
        )
    return wb


FAVORITE_HEADER = "favorite"


def _named_col(ws, header) -> int:
    """1-based column index of `header`, appended after the last column if missing."""
    headers = [c.value for c in ws[1]]
    if header in headers:
        return headers.index(header) + 1
    col = len([h for h in headers if h is not None]) + 1
    ws.cell(row=1, column=col, value=header)
    return col


def _favorite_col(ws) -> int:
    return _named_col(ws, FAVORITE_HEADER)


def ensure_favorite_column() -> None:
    wb = _load_workbook()
    ws = wb.active
    if FAVORITE_HEADER not in [c.value for c in ws[1]]:
        ws.cell(row=1, column=len(EXPECTED_HEADERS) + 1, value=FAVORITE_HEADER)
        wb.save(SESSIONS_XLSX)


def set_favorite(row_index: int, value: bool) -> None:
    wb = _load_workbook()
    ws = wb.active
    ws.cell(row=row_index, column=_favorite_col(ws), value=bool(value))
    wb.save(SESSIONS_XLSX)


GESTURE_SET_HEADER = "gesture_set"


def ensure_gesture_set_column() -> None:
    wb = _load_workbook()
    ws = wb.active
    if GESTURE_SET_HEADER not in [c.value for c in ws[1]]:
        _named_col(ws, GESTURE_SET_HEADER)
        wb.save(SESSIONS_XLSX)


def _set_from_model_name(mp):
    base = os.path.basename(str(mp or "")).lower()
    for s in ("rps", "6cl", "4cl"):
        if f"_{s}_" in base or f"_{s}." in base:
            return s
    return None


def get_gesture_set(row_index: int):
    """Gesture set of a session row: the stored column if present, else
    inferred from the model filename, else from n_classes (4->4cl, 6->6cl)."""
    wb = _load_workbook()
    ws = wb.active
    if row_index < 2 or row_index > ws.max_row:
        return None
    headers = [c.value for c in ws[1]]
    if GESTURE_SET_HEADER in headers:
        v = ws.cell(row=row_index,
                    column=headers.index(GESTURE_SET_HEADER) + 1).value
        if v:
            return str(v).strip()
    s = _set_from_model_name(ws.cell(row=row_index, column=COL["model_path"]).value)
    if s:
        return s
    nc = ws.cell(row=row_index, column=COL["n_classes"]).value
    if nc == 6:
        return "6cl"
    if nc == 4:
        return "4cl"
    return None


# ---------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------
def add_session(
    subject_id: str,
    n_classes: int,
    takes_batch1: int = 0,
    takes_batch2: int = 0,
    acc_offline: Optional[float] = None,
    f1_offline:  Optional[float] = None,
    model_path:  Optional[str]   = None,
    source_log:  Optional[str]   = None,
    data_folder: Optional[str]   = None,
    gesture_set: Optional[str]   = None,
    when: Optional[datetime]     = None,
) -> int:
    """Append a new session row.

    Returns the 1-based row index of the new row (useful so the caller
    can later update_online_metrics with the same row number).

    `when` defaults to datetime.now(). It is split into date (YYYY-MM-DD)
    and hour (HH:MM:SS) columns for human readability in Excel.

    acc_online / f1_online are intentionally left blank — they will be
    filled later by Phase 5 via update_online_metrics().
    """
    when = when or datetime.now()
    wb = _load_workbook()
    ws = wb.active

    new_row = [
        subject_id,
        when.strftime("%Y-%m-%d"),
        when.strftime("%H:%M:%S"),
        int(n_classes),
        int(takes_batch1),
        int(takes_batch2),
        acc_offline if acc_offline is not None else "",
        f1_offline  if f1_offline  is not None else "",
        "",  # acc_online — to be filled by Phase 5
        "",  # f1_online  — to be filled by Phase 5
        model_path or "",
        source_log or "",
        data_folder or "",
    ]
    ws.append(new_row)
    new_index = ws.max_row
    if gesture_set:
        _named_col(ws, FAVORITE_HEADER)  # keep favorite before gesture_set
        ws.cell(row=new_index,
                column=_named_col(ws, GESTURE_SET_HEADER), value=str(gesture_set))
    wb.save(SESSIONS_XLSX)

    # ws.max_row points to the newly appended row
    return ws.max_row


def update_online_metrics(row_index: int,
                           acc_online: float,
                           f1_online: float) -> None:
    """Write acc_online and f1_online into the row at `row_index`.

    `row_index` must come from a previous add_session() call (or from
    list_sessions_for_subject which returns it as 'row_index').
    """
    wb = _load_workbook()
    ws = wb.active

    if row_index < 2 or row_index > ws.max_row:
        raise IndexError(
            f"row_index={row_index} out of range (sheet has {ws.max_row} rows)"
        )

    ws.cell(row=row_index, column=COL["acc_online"], value=float(acc_online))
    ws.cell(row=row_index, column=COL["f1_online"],  value=float(f1_online))
    wb.save(SESSIONS_XLSX)


def get_subject_id(row_index: int):
    wb = _load_workbook()
    ws = wb.active
    if row_index < 2 or row_index > ws.max_row:
        return None
    v = ws.cell(row=row_index, column=COL["subject_id"]).value
    return str(v).strip() if v is not None else None


def find_row_by_model_path(model_path):
    """Excel row_index whose model_path matches (by basename); most recent wins."""
    if not model_path:
        return None
    target = os.path.basename(str(model_path)).strip().lower()
    wb = _load_workbook()
    ws = wb.active
    col = COL["model_path"]
    found = None
    for r in range(2, ws.max_row + 1):
        mp = ws.cell(row=r, column=col).value
        if mp and os.path.basename(str(mp)).strip().lower() == target:
            found = r
    return found


def list_sessions_for_subject(subject_id: str) -> list[dict]:
    """Return all sessions belonging to a subject, oldest first.

    Each dict carries 'row_index' (1-based, matches the Excel row) in
    addition to the regular columns, so the caller can update it later.
    """
    wb = _load_workbook()
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        if str(row[0]).strip() != subject_id:
            continue
        d = {h: row[idx] for idx, h in enumerate(EXPECTED_HEADERS)}
        d["row_index"] = i
        out.append(d)
    return out


def list_all_sessions() -> list[dict]:
    """Return every session row. Useful for a global overview."""
    wb = _load_workbook()
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        d = {h: row[idx] for idx, h in enumerate(EXPECTED_HEADERS)}
        d["row_index"] = i
        out.append(d)
    return out


def count_sessions_for_subject(subject_id: str) -> int:
    """Number of session rows belonging to a subject."""
    return len(list_sessions_for_subject(subject_id))


def delete_sessions_for_subjects(subject_ids) -> int:
    """Delete every session row whose subject_id is in `subject_ids`.

    Accepts a single id (str) or an iterable of ids. Returns the number
    of rows actually removed.
    """
    if isinstance(subject_ids, str):
        subject_ids = [subject_ids]
    targets = {str(s).strip() for s in subject_ids if str(s).strip()}
    if not targets:
        return 0

    wb = _load_workbook()
    ws = wb.active

    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        if str(row[0]).strip() in targets:
            rows_to_delete.append(i)

    for i in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(i, 1)

    if rows_to_delete:
        wb.save(SESSIONS_XLSX)
    return len(rows_to_delete)


# ---------------------------------------------------------------------
# CLI smoke test
# ---------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    import json

    cmd = sys.argv[1] if len(sys.argv) > 1 else "list"

    if cmd == "list":
        for s in list_all_sessions():
            print(json.dumps(s, default=str))

    elif cmd == "add":
        if len(sys.argv) < 4:
            print("usage: sessions_registry.py add <subject_id> <n_classes> "
                  "[acc_offline]")
            sys.exit(2)
        sid = sys.argv[2]
        ncl = int(sys.argv[3])
        acc = float(sys.argv[4]) if len(sys.argv) > 4 else None
        row = add_session(sid, ncl, acc_offline=acc)
        print(f"added row {row}")

    elif cmd == "for":
        if len(sys.argv) < 3:
            print("usage: sessions_registry.py for <subject_id>")
            sys.exit(2)
        for s in list_sessions_for_subject(sys.argv[2]):
            print(json.dumps(s, default=str))

    else:
        print(f"unknown command: {cmd}")
        print("commands: list | add <subject_id> <n_classes> [acc] "
              "| for <subject_id>")
        sys.exit(2)


def delete_session_by_match(subject_id, date, hour, model_path=""):
    """Delete the first row matching (subject_id, date, hour, model_path).

    Exact string match on the four fields. Returns 1 if a row was deleted,
    0 if nothing matched.
    """
    subject_id = str(subject_id or "").strip()
    date = str(date or "").strip()
    hour = str(hour or "").strip()
    model_path = str(model_path or "").strip()
    wb = _load_workbook(); ws = wb.active
    c_sub, c_date = COL["subject_id"], COL["date"]
    c_hour, c_model = COL["hour"], COL["model_path"]
    target = None
    for i in range(2, ws.max_row + 1):
        if (str(ws.cell(row=i, column=c_sub).value or "").strip() == subject_id
                and str(ws.cell(row=i, column=c_date).value or "").strip() == date
                and str(ws.cell(row=i, column=c_hour).value or "").strip() == hour
                and str(ws.cell(row=i, column=c_model).value or "").strip() == model_path):
            target = i; break
    if target is None:
        return 0
    ws.delete_rows(target, 1); wb.save(SESSIONS_XLSX)
    return 1